import os
import smtplib
from datetime import date
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from state import AnalysisResult, Recommendation, ResearchResult, State, StockData

GMAIL_SENDER = os.getenv("GMAIL_SENDER", "")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")
GMAIL_RECIPIENT = os.getenv("GMAIL_RECIPIENT", "")

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F4E79")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

SUMMARY_HEADERS = [
    "Ticker",
    "Company",
    "Price ($)",
    "Change ($)",
    "Change (%)",
    "Volume",
    "Market Cap",
    "P/E Ratio",
    "Earnings Date",
    "Sentiment",
    "Recommendation",
    "Confidence",
    "Reasoning",
]

RAW_HEADERS = [
    "Ticker",
    "Company",
    "Price",
    "Change",
    "Change_Pct",
    "Volume",
    "Avg_Vol_3M",
    "Market_Cap",
    "PE_Ratio",
    "Earnings_Date",
    "Week_52_Change_Pct",
    "Week_52_Low",
    "Week_52_High",
    "News_Summary",
    "Key_Events",
    "Technical_Analysis",
    "Sentiment",
    "Action",
    "Reasoning",
    "Confidence",
]


def _find_top_gainer(stocks: list[StockData]) -> StockData:
    return max(stocks, key=lambda s: s.change_percent)


def _find_top_loser(stocks: list[StockData]) -> StockData:
    return min(stocks, key=lambda s: s.change_percent)


def _find_top_recommended(recommendations: list[Recommendation]) -> list[Recommendation]:
    buys = sorted(
        [r for r in recommendations if r.action == "Buy"],
        key=lambda r: r.confidence,
        reverse=True,
    )
    if len(buys) >= 3:
        return buys[:3]
    holds = sorted(
        [r for r in recommendations if r.action == "Hold"],
        key=lambda r: r.confidence,
        reverse=True,
    )
    return (buys + holds)[:3]


def _style_header_row(ws: Worksheet, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def _build_summary_sheet(
    ws: Worksheet,
    stocks: list[StockData],
    analyses: list[AnalysisResult],
    recommendations: list[Recommendation],
) -> None:
    today = date.today().strftime("%B %d, %Y")

    ws.cell(row=1, column=1, value=f"Daily Movers Report – {today}")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = LEFT

    for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(row=2, column=col_idx, value=header)
    _style_header_row(ws, 2, len(SUMMARY_HEADERS))

    analysis_map = {a.ticker: a for a in analyses}
    rec_map = {r.ticker: r for r in recommendations}

    top_gainer_ticker = _find_top_gainer(stocks).ticker
    top_loser_ticker = _find_top_loser(stocks).ticker
    top_rec_tickers = {r.ticker for r in _find_top_recommended(recommendations)}

    for row_offset, stock in enumerate(stocks):
        row = 3 + row_offset
        ticker = stock.ticker
        analysis = analysis_map.get(ticker)
        rec = rec_map.get(ticker)

        values = [
            ticker,
            stock.company_name,
            stock.price,
            stock.change,
            stock.change_percent,
            stock.volume,
            stock.market_cap,
            stock.pe_ratio if stock.pe_ratio is not None else "—",
            stock.earnings_date if stock.earnings_date is not None else "—",
            analysis.sentiment if analysis else "—",
            rec.action if rec else "—",
            rec.confidence if rec else "—",
            rec.reasoning if rec else "—",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            if col_idx in (3, 4, 5, 6, 8, 9, 12):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

        if ticker == top_gainer_ticker:
            fill = GREEN_FILL
        elif ticker == top_loser_ticker:
            fill = RED_FILL
        elif ticker in top_rec_tickers:
            fill = GOLD_FILL
        else:
            fill = None

        if fill:
            for col_idx in range(1, len(SUMMARY_HEADERS) + 1):
                ws.cell(row=row, column=col_idx).fill = fill

    widths = [10, 22, 12, 12, 12, 16, 14, 12, 14, 12, 16, 12, 40]
    for i, w in enumerate(widths, start=1):
        col_letter = chr(64 + i)
        ws.column_dimensions[col_letter].width = w


def _build_raw_sheet(
    ws: Worksheet,
    stocks: list[StockData],
    analyses: list[AnalysisResult],
    recommendations: list[Recommendation],
    research_results: list[ResearchResult],
) -> None:
    for col_idx, header in enumerate(RAW_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    _style_header_row(ws, 1, len(RAW_HEADERS))

    analysis_map = {a.ticker: a for a in analyses}
    rec_map = {r.ticker: r for r in recommendations}
    research_map = {r.ticker: r for r in research_results}

    for row_offset, stock in enumerate(stocks):
        row = 2 + row_offset
        ticker = stock.ticker
        analysis = analysis_map.get(ticker)
        rec = rec_map.get(ticker)
        research = research_map.get(ticker)

        values = [
            ticker,
            stock.company_name,
            stock.price,
            stock.change,
            stock.change_percent,
            stock.volume,
            stock.avg_volume_3m,
            stock.market_cap,
            stock.pe_ratio,
            stock.earnings_date,
            stock.week_52_change_pct,
            stock.week_52_low,
            stock.week_52_high,
            research.news_summary if research else "",
            "; ".join(research.key_events) if research else "",
            analysis.technical_analysis if analysis else "",
            analysis.sentiment if analysis else "",
            rec.action if rec else "",
            rec.reasoning if rec else "",
            rec.confidence if rec else "",
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=val)

    for col_idx in range(1, len(RAW_HEADERS) + 1):
        col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"
        ws.column_dimensions[col_letter].width = 22


def _send_email(subject: str, body: str, excel_path: str) -> None:
    if not all([GMAIL_SENDER, GMAIL_APP_PASSWORD, GMAIL_RECIPIENT]):
        return

    msg = MIMEMultipart()
    msg["From"] = GMAIL_SENDER
    msg["To"] = GMAIL_RECIPIENT
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(excel_path, "rb") as f:
        attachment = MIMEBase("application", "octet-stream")
        attachment.set_payload(f.read())
        encoders.encode_base64(attachment)
        attachment.add_header(
            "Content-Disposition", "attachment", filename=excel_path
        )
        msg.attach(attachment)

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(GMAIL_SENDER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_SENDER, GMAIL_RECIPIENT, msg.as_string())


async def generate_report_node(state: State) -> State:
    wb = Workbook()

    ws_summary = wb.active
    assert ws_summary is not None
    ws_summary.title = "Daily Summary"
    _build_summary_sheet(
        ws_summary,
        state.stocks,
        state.analysis_results,
        state.recommendations,
    )

    ws_raw = wb.create_sheet("Raw Data")
    _build_raw_sheet(
        ws_raw,
        state.stocks,
        state.analysis_results,
        state.recommendations,
        state.research_results,
    )

    filename = f"daily_movers_report_{date.today().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)

    return state.model_copy(update={"excel_path": filename})


async def generate_email_node(state: State) -> State:
    stocks = state.stocks
    recommendations = state.recommendations

    top_gainer = _find_top_gainer(stocks)
    top_loser = _find_top_loser(stocks)
    top_recs = _find_top_recommended(recommendations)

    rec_map = {r.ticker: r for r in recommendations}

    lines: list[str] = []
    lines.append(f"Daily Movers Report – {date.today().strftime('%B %d, %Y')}")
    lines.append("=" * 55)
    lines.append("")

    lines.append("TOP GAINER")
    lines.append("-" * 30)
    tg_rec = rec_map.get(top_gainer.ticker)
    lines.append(f"  {top_gainer.ticker} ({top_gainer.company_name})")
    lines.append(f"  Price: ${top_gainer.price:.2f}  |  Change: +{top_gainer.change_percent}%")
    lines.append(f"  Recommendation: {tg_rec.action if tg_rec else '—'}")
    lines.append("")

    lines.append("TOP LOSER")
    lines.append("-" * 30)
    tl_rec = rec_map.get(top_loser.ticker)
    lines.append(f"  {top_loser.ticker} ({top_loser.company_name})")
    lines.append(f"  Price: ${top_loser.price:.2f}  |  Change: {top_loser.change_percent}%")
    lines.append(f"  Recommendation: {tl_rec.action if tl_rec else '—'}")
    lines.append("")

    lines.append("TOP 3 RECOMMENDATIONS")
    lines.append("-" * 30)
    for i, rec in enumerate(top_recs, start=1):
        lines.append(
            f"  {i}. {rec.ticker} – {rec.action} "
            f"(confidence: {rec.confidence:.0%})"
        )
        lines.append(f"     {rec.reasoning}")
    lines.append("")
    lines.append("-" * 55)
    lines.append(f"Full details in: {state.excel_path or 'report not generated'}")

    email_text = "\n".join(lines)

    summary_filename = f"daily_movers_summary_{date.today().strftime('%Y%m%d')}.txt"
    with open(summary_filename, "w") as f:
        f.write(email_text)

    if state.excel_path:
        subject = f"Daily Movers Report – {date.today().strftime('%B %d, %Y')}"
        _send_email(subject, email_text, state.excel_path)

    return state.model_copy(update={"email_summary": email_text})
